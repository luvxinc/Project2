"""
VMA Training SOP Seed Script
Import SOP data from IR-2004-001.xlsx into vma_training_sops table.

Usage: python3 scripts/seed-training-sops.py
"""
import openpyxl
import psycopg2
import uuid
from datetime import datetime

# Database connection
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'dbname': 'mgmt_v2',
    'user': 'aaron',
    'password': '***REDACTED_PASSWORD***',
}

EXCEL_PATH = '/Users/aaron/Dropbox/VMA/Documents_Technical/IR - Internal Reports/IR-2004-XXX LCD List of Controlled Documents/IR-2004-001.xlsx'

DEFAULT_DATE = datetime(2025, 6, 1)  # Default for empty implementation dates

def main():
    # Read Excel
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    records = []
    for row_idx in range(3, ws.max_row + 1):
        no = ws.cell(row=row_idx, column=1).value       # A: No.
        doc_no = ws.cell(row=row_idx, column=2).value    # B: Document No
        doc_name = ws.cell(row=row_idx, column=3).value  # C: Document name
        rev = ws.cell(row=row_idx, column=4).value       # D: Rev.
        impl_date = ws.cell(row=row_idx, column=5).value # E: Implementation date
        struct_class = ws.cell(row=row_idx, column=8).value  # H: Structure Classification
        doc_type = ws.cell(row=row_idx, column=9).value      # I: Document Type

        # Skip empty rows and non-data rows (e.g. "Updated by/Date:")
        if not doc_no or not doc_name:
            continue
        if not isinstance(no, (int, float)):
            continue
        if not rev or not struct_class or not doc_type:
            print(f"  SKIP Row {row_idx}: missing required fields (no={no}, doc_no={doc_no})")
            continue

        # Parse date
        if isinstance(impl_date, datetime):
            eff_date = impl_date
        elif impl_date:
            try:
                eff_date = datetime.strptime(str(impl_date), '%Y-%m-%d')
            except:
                eff_date = DEFAULT_DATE
                print(f"  FALLBACK date for Row {row_idx}: {doc_no} (was: {impl_date})")
        else:
            eff_date = DEFAULT_DATE

        records.append({
            'id': str(uuid.uuid4()),
            'seq_no': int(no),
            'sop_no': str(doc_no).strip(),
            'name': str(doc_name).strip(),
            'version': str(rev).strip(),
            'effective_date': eff_date,
            'structure_classification': str(struct_class).strip(),
            'document_type': str(doc_type).strip(),
            'training_required': True,
            'status': 'ACTIVE',
        })

    print(f"\n{'='*60}")
    print(f"Records parsed: {len(records)}")
    print(f"{'='*60}")

    # Show summary by document type
    type_counts = {}
    for r in records:
        dt = r['document_type']
        type_counts[dt] = type_counts.get(dt, 0) + 1
    print("\nBy Document Type:")
    for dt, count in sorted(type_counts.items()):
        print(f"  {dt}: {count}")

    # Show summary by structure classification
    class_counts = {}
    for r in records:
        sc = r['structure_classification']
        class_counts[sc] = class_counts.get(sc, 0) + 1
    print("\nBy Structure Classification:")
    for sc, count in sorted(class_counts.items()):
        print(f"  {sc}: {count}")

    # Count records with default date
    default_date_count = sum(1 for r in records if r['effective_date'] == DEFAULT_DATE)
    print(f"\nRecords with default date (2025-06-01): {default_date_count}")

    # Insert into database
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Clear existing data
    cur.execute("DELETE FROM vma_training_sops")
    deleted = cur.rowcount
    print(f"\nCleared {deleted} existing records")

    # Insert
    insert_sql = """
        INSERT INTO vma_training_sops (
            id, seq_no, sop_no, name, version,
            effective_date, structure_classification, document_type,
            training_required, status, created_at, updated_at
        ) VALUES (
            %s, %s, %s, %s, %s,
            %s, %s, %s,
            %s, %s, NOW(), NOW()
        )
    """

    inserted = 0
    for r in records:
        try:
            cur.execute(insert_sql, (
                r['id'], r['seq_no'], r['sop_no'], r['name'], r['version'],
                r['effective_date'], r['structure_classification'], r['document_type'],
                r['training_required'], r['status'],
            ))
            inserted += 1
        except Exception as e:
            print(f"  ERROR inserting #{r['seq_no']} {r['sop_no']}: {e}")
            conn.rollback()
            cur = conn.cursor()

    conn.commit()
    cur.close()
    conn.close()

    print(f"\n✅ Inserted {inserted}/{len(records)} records successfully")
    print(f"{'='*60}")

if __name__ == '__main__':
    main()
