"""
新建采购订单 - 模板相关API
包含：下载模板、解析Excel

[P0-2 Fix] 统一使用 hub.check_perm
[P1-2 Fix] 使用 logging 替代 traceback.print_exc
"""
import os
import logging
from datetime import datetime

from django.http import JsonResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.conf import settings as django_settings

from ..hub import check_perm
from django.utils.translation import gettext as _

logger = logging.getLogger(__name__)


def create_po_template(filepath):
    """创建采购订单Excel模板"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "采购订单明细"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ["SKU", "数量", "单价"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        ws.cell(row=2, column=1, value="SKU001")
        ws.cell(row=2, column=2, value=100)
        ws.cell(row=2, column=3, value=10.50)
        
        ws.cell(row=3, column=1, value="SKU002")
        ws.cell(row=3, column=2, value=50)
        ws.cell(row=3, column=3, value=25.00)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        
    except ImportError:
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'wb') as f:
            f.write(b'')
        raise Exception("openpyxl not installed, cannot create Excel template")


@login_required(login_url='web_ui:login')
@require_http_methods(["GET"])
def download_po_template_api(request):
    """
    下载采购订单Excel模板
    Permission: module.purchase.po.add
    """
    if not check_perm(request.user, 'module.purchase.po.add'):
        return JsonResponse({'error': _('无权限')}, status=403)
    
    project_root = os.path.dirname(django_settings.BASE_DIR)
    template_path = os.path.join(
        project_root, 
        'data', 'templates_csv', 'in_po_upload.xlsx'
    )
    
    if not os.path.exists(template_path):
        try:
            create_po_template(template_path)
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': _('模板文件不存在且创建失败: {error}').format(error=str(e))
            }, status=404)
    
    try:
        response = FileResponse(
            open(template_path, 'rb'),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="in_po_upload_template.xlsx"'
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required(login_url='web_ui:login')
@require_http_methods(["POST"])
def parse_po_excel_api(request):
    """
    解析上传的采购订单Excel文件
    Permission: module.purchase.po.add
    """
    if not check_perm(request.user, 'module.purchase.po.add'):
        return JsonResponse({'error': _('无权限')}, status=403)
    
    try:
        import openpyxl
        from io import BytesIO
        from difflib import SequenceMatcher
        
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({
                'success': False, 
                'error_type': 'no_file',
                'message': _('请上传文件')
            })
        
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'error_type': 'invalid_format',
                'message': _('仅支持 .xlsx 或 .xls 格式的Excel文件')
            })
        
        expected_supplier_code = request.POST.get('supplier_code', '').strip()
        expected_date = request.POST.get('po_date', '').strip()
        expected_currency = request.POST.get('currency', '').strip().upper()
        
        try:
            wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()), data_only=True)
            ws = wb.active
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error_type': 'parse_error',
                'message': _('无法解析Excel文件: {error}').format(error=str(e))
            })
        
        # 1. 验证B1格内容
        b1_value = ws['B1'].value
        expected_header = "Eaglestar Purchase Order Form"
        if not b1_value or str(b1_value).strip() != expected_header:
            return JsonResponse({
                'success': False,
                'error_type': 'header_mismatch',
                'message': _('文件格式错误：B1单元格应为 "{expected}"，实际为 "{actual}"').format(expected=expected_header, actual=b1_value)
            })
        
        # 2. 验证C2 (supplier_code) 和 E2 (日期)
        c2_value = ws['C2'].value
        e2_value = ws['E2'].value
        
        if not c2_value or str(c2_value).strip() != expected_supplier_code:
            return JsonResponse({
                'success': False,
                'error_type': 'supplier_mismatch',
                'message': _('供应商不匹配：文件中为 "{actual}"，期望为 "{expected}"').format(actual=c2_value, expected=expected_supplier_code)
            })
        
        file_date_str = None
        if e2_value:
            if isinstance(e2_value, datetime):
                file_date_str = e2_value.strftime('%Y-%m-%d')
            elif isinstance(e2_value, str):
                date_formats = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y', '%Y.%m.%d']
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(e2_value.strip(), fmt)
                        file_date_str = parsed_date.strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue
        
        if file_date_str != expected_date:
            return JsonResponse({
                'success': False,
                'error_type': 'date_mismatch',
                'requires_reupload': True,
                'message': _('日期不匹配：文件中为 "{actual_raw}" (解析为 {actual_parsed})，期望为 "{expected}"').format(actual_raw=e2_value, actual_parsed=file_date_str, expected=expected_date)
            })
        
        # 3. 验证G2 (货币)
        g2_value = ws['G2'].value
        file_currency = str(g2_value).strip().upper() if g2_value else ''
        if expected_currency and file_currency != expected_currency:
            return JsonResponse({
                'success': False,
                'error_type': 'currency_mismatch',
                'requires_reupload': True,
                'message': _('货币不匹配：文件中为 "{actual}"，期望为 "{expected}"').format(actual=g2_value, expected=expected_currency)
            })
        
        # 4. 获取有效SKU列表
        from core.components.db.client import DBClient
        from difflib import SequenceMatcher
        
        sku_df = DBClient.read_df("SELECT DISTINCT SKU FROM Data_COGS WHERE SKU IS NOT NULL")
        valid_skus = set(sku_df['SKU'].str.upper().tolist()) if not sku_df.empty else set()
        valid_skus_list = sorted(valid_skus)
        
        def get_sku_suggestions(bad_sku, limit=5):
            if not bad_sku:
                return []
            bad_sku_upper = bad_sku.upper()
            scored = []
            for sku in valid_skus:
                score = SequenceMatcher(None, bad_sku_upper, sku).ratio()
                if score > 0.4:
                    scored.append((sku, score))
            scored.sort(key=lambda x: x[1], reverse=True)
            return [s[0] for s in scored[:limit]]
        
        # 5. 解析商品数据 (从B5开始)
        items = []
        sku_errors = []
        data_errors = []
        
        for row_num in range(5, ws.max_row + 1):
            sku_raw = ws.cell(row=row_num, column=2).value  # B列
            qty = ws.cell(row=row_num, column=3).value      # C列
            price = ws.cell(row=row_num, column=4).value    # D列
            
            # 跳过空行
            if not sku_raw and not qty and not price:
                continue
            
            sku = str(sku_raw).strip().upper() if sku_raw else ''
            
            row_item = {
                'row': row_num,
                'sku': sku,
                'sku_original': str(sku_raw).strip() if sku_raw else '',
                'qty': None,
                'unit_price': None,
                'sku_error': None,
                'qty_error': None,
                'price_error': None,
                'suggestions': []
            }
            
            # 验证SKU
            if not sku:
                row_item['sku_error'] = _('SKU不能为空')
            elif sku not in valid_skus:
                row_item['sku_error'] = _('SKU不存在')
                row_item['suggestions'] = get_sku_suggestions(sku)
            
            # 验证数量
            try:
                qty_val = int(qty) if qty is not None else None
                if qty_val is None:
                    row_item['qty_error'] = _('数量不能为空')
                elif qty_val <= 0:
                    row_item['qty_error'] = _('数量必须大于0')
                row_item['qty'] = qty_val
            except (ValueError, TypeError):
                row_item['qty_error'] = _('数量必须为整数')
                row_item['qty'] = None
            
            # 验证单价
            try:
                if price is not None:
                    price_val = float(price)
                    if price_val <= 0:
                        row_item['price_error'] = _('单价必须大于0')
                    row_item['unit_price'] = round(price_val, 5)
                else:
                    row_item['price_error'] = _('单价不能为空')
                    row_item['unit_price'] = None
            except (ValueError, TypeError):
                row_item['price_error'] = _('单价必须为数字')
                row_item['unit_price'] = None
            
            if row_item['sku_error']:
                sku_errors.append(row_item)
            elif row_item['qty_error'] or row_item['price_error']:
                data_errors.append(row_item)
            
            items.append(row_item)
        
        if not items:
            return JsonResponse({
                'success': False,
                'error_type': 'no_data',
                'message': _('未在Excel中找到商品数据（从B5行开始）')
            })
        
        valid_items = [item for item in items if not item['sku_error'] and not item['qty_error'] and not item['price_error']]
        
        # 如果有SKU错误，返回修正界面数据
        if sku_errors:
            return JsonResponse({
                'success': False,
                'error_type': 'sku_errors',
                'requires_reupload': False,
                'can_fix': True,
                'message': _('发现 {n} 行SKU错误，请在线修正').format(n=len(sku_errors)),
                'items': [{'row': item['row'], 'sku': item['sku'], 'sku_original': item['sku_original'], 'qty': item['qty'], 'unit_price': item['unit_price'], 'sku_error': item['sku_error'], 'qty_error': item['qty_error'], 'price_error': item['price_error']} for item in items],
                'sku_errors': [{'row': item['row'], 'sku': item['sku'], 'sku_original': item['sku_original'], 'error': item['sku_error'], 'suggestions': item.get('suggestions', []), 'qty': item['qty'], 'unit_price': item['unit_price']} for item in sku_errors],
                'data_errors': [{'row': item['row'], 'sku': item['sku'], 'qty': item['qty'], 'qty_error': item['qty_error'], 'unit_price': item['unit_price'], 'price_error': item['price_error']} for item in data_errors],
                'valid_count': len(valid_items),
                'all_skus': valid_skus_list
            })
        
        # 如果有数据错误
        if data_errors:
            return JsonResponse({
                'success': False,
                'error_type': 'data_errors',
                'requires_reupload': False,
                'can_fix': True,
                'message': _('发现 {n} 行数据错误，请在线修正').format(n=len(data_errors)),
                'items': [{'row': item['row'], 'sku': item['sku'], 'qty': item['qty'], 'unit_price': item['unit_price'], 'qty_error': item['qty_error'], 'price_error': item['price_error']} for item in items],
                'data_errors': [{'row': item['row'], 'sku': item['sku'], 'qty': item['qty'], 'qty_error': item['qty_error'], 'unit_price': item['unit_price'], 'price_error': item['price_error']} for item in data_errors],
                'valid_count': len(valid_items)
            })
        
        # 全部验证通过
        return JsonResponse({
            'success': True,
            'message': _('成功解析 {n} 条商品数据').format(n=len(valid_items)),
            'items': [{'row': item['row'], 'sku': item['sku'], 'qty': item['qty'], 'unit_price': item['unit_price']} for item in valid_items],
            'item_count': len(valid_items)
        })


        
    except ImportError as ie:
        logger.exception("依赖模块导入失败")
        return JsonResponse({
            'success': False,
            'error_type': 'dependency_error',
            'message': _('依赖模块导入失败: {error}').format(error=str(ie))
        }, status=500)
    except Exception as e:
        logger.exception("服务器错误")
        return JsonResponse({
            'success': False,
            'error_type': 'server_error',
            'message': _('服务器错误: {error}').format(error=str(e))
        }, status=500)


@login_required(login_url='web_ui:login')
@require_http_methods(["POST"])
def generate_prefilled_template_api(request):
    """
    生成预填并锁定的采购订单模板
    - 读取原模板 in_po_upload.xlsx
    - 写入用户选择的日期到 E2
    - 写入用户选择的供应商到 C2
    - 锁定除 B5:D∞ 和 G2 外的所有单元格（密码1522）
    - 返回生成的预填模板（不修改原模板文件）
    Permission: module.purchase.po.add
    """
    import json
    from io import BytesIO
    
    if not check_perm(request.user, 'module.purchase.po.add'):
        return JsonResponse({'error': _('无权限')}, status=403)
    
    try:
        import openpyxl
        from openpyxl.worksheet.protection import SheetProtection
        
        # 获取参数
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = request.POST
        
        po_date = data.get('po_date', '').strip()
        supplier_code = data.get('supplier_code', '').strip()
        
        if not po_date or not supplier_code:
            return JsonResponse({
                'success': False,
                'message': _('请提供订单日期和供应商')
            }, status=400)
        
        # 读取原模板
        project_root = os.path.dirname(django_settings.BASE_DIR)
        template_path = os.path.join(
            project_root, 
            'data', 'templates_csv', 'in_po_upload.xlsx'
        )
        
        if not os.path.exists(template_path):
            return JsonResponse({
                'success': False,
                'message': _('模板文件不存在')
            }, status=404)
        
        # 加载工作簿
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 写入日期到 E2 (仅年月日)
        ws['E2'] = po_date
        
        # 写入供应商到 C2
        ws['C2'] = supplier_code
        
        # ========== 锁定逻辑 ==========
        # 参照 send_create/template.py 的实现方式
        from openpyxl.styles import Protection
        
        # 解锁 G2（备注字段）
        ws['G2'].protection = Protection(locked=False)
        
        # 解锁 B、C、D 列从第5行开始到第1000行
        # 这覆盖用户可能需要输入商品的区域
        for row_num in range(5, 1001):
            ws.cell(row=row_num, column=2).protection = Protection(locked=False)  # B列
            ws.cell(row=row_num, column=3).protection = Protection(locked=False)  # C列
            ws.cell(row=row_num, column=4).protection = Protection(locked=False)  # D列
        
        # 启用工作表保护，密码为1522
        ws.protection.sheet = True
        ws.protection.password = '1522'
        ws.protection.enable()

        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 返回文件
        from django.http import HttpResponse
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'PO_Template_{supplier_code}_{po_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except ImportError as ie:
        logger.exception("依赖模块导入失败")
        return JsonResponse({
            'success': False,
            'message': _('依赖模块导入失败: {error}').format(error=str(ie))
        }, status=500)
    except Exception as e:
        logger.exception("生成预填模板失败")
        return JsonResponse({
            'success': False,
            'message': _('生成预填模板失败: {error}').format(error=str(e))
        }, status=500)

