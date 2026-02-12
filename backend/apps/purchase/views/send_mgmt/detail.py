"""
发货单管理 - 发货单详情和下载
"""
import logging

import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.conf import settings as django_settings
from django.http import JsonResponse, FileResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET

from django.utils.translation import gettext as _
import json

from ..hub import check_perm
from core.components.db.client import DBClient

logger = logging.getLogger(__name__)


@login_required(login_url='web_ui:login')
@require_GET
def po_detail_api(request):
    """
    获取单个发货单详情
    URL: /dashboard/purchase/api/send_mgmt/detail/?logistic_num=xxx
    
    返回:
    1. base: 物流单基础信息 (物流单号, 发货日期)
    2. logistics: 物流单详细信息 (日期, 托盘, 重量, 单价, 总价, 汇率, 版本, 操作人等)
    3. detail: 物流单货物信息 (版本, 操作人, 日期, 备注, 货物价值)
    """
    if not check_perm(request.user, 'module.purchase.send.mgmt'):
        return JsonResponse({'success': False, 'message': 'Forbidden'}, status=403)
    
    logistic_num = request.GET.get('logistic_num', '').strip()
    if not logistic_num:
        return JsonResponse({'success': False, 'message': _('缺少物流单号')}, status=400)
    
    try:
        # 1. 获取物流单基础信息 (in_send 表中 seq 最大的记录)
        base_df = DBClient.read_df("""
            SELECT 
                logistic_num,
                date_sent,
                date_eta,
                pallets,
                total_weight,
                price_kg,
                total_price,
                usd_rmb,
                mode,
                seq,
                `by`,
                date_record,
                note,
                CAST(SUBSTRING(seq, 2) AS UNSIGNED) as seq_num
            FROM in_send 
            WHERE logistic_num = :logistic_num
            ORDER BY seq_num DESC
            LIMIT 1
        """, {'logistic_num': logistic_num})
        
        if base_df.empty:
            return JsonResponse({'success': False, 'message': _('发货单不存在')}, status=404)
        
        row = base_df.iloc[0]
        
        # 基础信息: 物流单号, 发货日期
        base_info = {
            'logistic_num': logistic_num,
            'date_sent': str(row['date_sent']) if row['date_sent'] else '-'
        }
        
        # 物流单详细信息 (来自 in_send 表最大 seq 的记录)
        mode_display = _('自动获取') if row['mode'] == 'A' else _('手动填写') if row['mode'] == 'M' else '-'
        logistics_info = {
            # 第一行
            'date_eta': str(row['date_eta']) if row['date_eta'] else '-',
            'pallets': int(row['pallets']) if row['pallets'] else 0,
            'total_weight': float(row['total_weight']) if row['total_weight'] else 0.0,
            # 第二行
            'price_kg': float(row['price_kg']) if row['price_kg'] else 0.0,
            'total_price': float(row['total_price']) if row['total_price'] else 0.0,
            'usd_rmb': float(row['usd_rmb']) if row['usd_rmb'] else 0.0,
            'mode': mode_display,
            # 第三行
            'seq': row['seq'] or '-',
            'by': row['by'] or '-',
            'date_record': str(row['date_record']) if row['date_record'] else '-',
            # 第四行
            'note': row['note'] or '-'
        }
        
        # 2. 获取货物明细信息 (in_send_list 表中 seq 最大的记录)
        detail_df = DBClient.read_df("""
            SELECT 
                seq,
                `by`,
                date,
                note,
                CAST(SUBSTRING(seq, 2) AS UNSIGNED) as seq_num
            FROM in_send_list 
            WHERE logistic_num = :logistic_num
            ORDER BY seq_num DESC
            LIMIT 1
        """, {'logistic_num': logistic_num})
        
        if not detail_df.empty:
            detail_row = detail_df.iloc[0]
            detail_info = {
                'seq': detail_row['seq'] or '-',
                'by': detail_row['by'] or '-',
                'date': str(detail_row['date']) if detail_row['date'] else '-',
                'note': detail_row['note'] or '-'
            }
        else:
            detail_info = {
                'seq': '-',
                'by': '-',
                'date': '-',
                'note': '-'
            }
        
        # 3. 计算货物价值 (in_send_final 中 sent_quantity * po_price 的和)
        cargo_df = DBClient.read_df("""
            SELECT SUM(sent_quantity * po_price) as cargo_value
            FROM in_send_final
            WHERE sent_logistic_num = :logistic_num
        """, {'logistic_num': logistic_num})
        
        cargo_value = 0.0
        if not cargo_df.empty and cargo_df.iloc[0]['cargo_value']:
            cargo_value = float(cargo_df.iloc[0]['cargo_value'])
        
        detail_info['cargo_value'] = cargo_value
        
        # 4. 获取货物列表 (in_send_final 中的记录)
        items_df = DBClient.read_df("""
            SELECT 
                po_num,
                po_sku,
                sent_quantity,
                po_price
            FROM in_send_final
            WHERE sent_logistic_num = :logistic_num
            ORDER BY po_num, po_sku
        """, {'logistic_num': logistic_num})
        
        items = []
        if not items_df.empty:
            # 获取所有涉及的订单号
            po_nums = list(items_df['po_num'].unique())
            
            # 获取每个订单的货币和汇率信息（从最新策略）
            currency_map = {}
            if po_nums:
                strategy_sql = """
                SELECT 
                    po_num,
                    (SELECT cur_currency FROM in_po_strategy s2 
                     WHERE s2.po_num = in_po_strategy.po_num 
                     ORDER BY date DESC, seq DESC LIMIT 1) as currency,
                    (SELECT cur_usd_rmb FROM in_po_strategy s3 
                     WHERE s3.po_num = in_po_strategy.po_num 
                     ORDER BY date DESC, seq DESC LIMIT 1) as usd_rmb
                FROM in_po_strategy
                WHERE po_num IN :po_nums
                GROUP BY po_num
                """
                strategy_df = DBClient.read_df(strategy_sql, {'po_nums': tuple(po_nums)})
                for _idx, srow in strategy_df.iterrows():
                    currency_map[srow['po_num']] = {
                        'currency': srow.get('currency') or 'RMB',
                        'usd_rmb': float(srow['usd_rmb']) if srow.get('usd_rmb') else 7.0
                    }
            
            for _idx, item_row in items_df.iterrows():
                po_num = item_row['po_num'] or '-'
                unit_price = float(item_row['po_price']) if item_row['po_price'] else 0.0
                quantity = int(item_row['sent_quantity']) if item_row['sent_quantity'] else 0
                
                # 获取该订单的货币和汇率
                po_strategy = currency_map.get(po_num, {'currency': 'RMB', 'usd_rmb': 7.0})
                currency = po_strategy['currency']
                usd_rmb = po_strategy['usd_rmb']
                
                # 计算双货币价值
                item_value = unit_price * quantity
                if currency == 'USD':
                    value_usd = round(item_value, 5)
                    value_rmb = round(item_value * usd_rmb, 5)
                else:
                    value_rmb = round(item_value, 5)
                    value_usd = round(item_value / usd_rmb, 5) if usd_rmb > 0 else 0
                
                items.append({
                    'po_num': po_num,
                    'sku': item_row['po_sku'] or '-',
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'currency': currency,
                    'usd_rmb': usd_rmb,
                    'value_rmb': value_rmb,
                    'value_usd': value_usd
                })
        
        detail_info['items'] = items
        
        # 计算货物价值汇总（双货币）
        total_cargo_rmb = sum(item['value_rmb'] for item in items)
        total_cargo_usd = sum(item['value_usd'] for item in items)
        detail_info['cargo_value_rmb'] = round(total_cargo_rmb, 5)
        detail_info['cargo_value_usd'] = round(total_cargo_usd, 5)
        
        return JsonResponse({
            'success': True,
            'data': {
                'base': base_info,
                'logistics': logistics_info,
                'detail': detail_info
            }
        })
        
    except Exception as e:
        logger.exception("操作失败")
        return JsonResponse({
            'success': False,
            'message': _('获取发货单详情失败: {error}').format(error=str(e))
        }, status=500)


@login_required(login_url='web_ui:login')
@require_GET
def download_po_excel_api(request):
    """
    下载发货单Excel文件(管理部门用)
    URL: /dashboard/purchase/api/send_mgmt/download/?logistic_num=xxx&type=mgmt
    
    使用模板 data/templates_csv/in_send_output_mgmt.xlsx 生成Excel文件
    """
    if not check_perm(request.user, 'module.purchase.send.mgmt'):
        return JsonResponse({'success': False, 'message': 'Forbidden'}, status=403)
    
    logistic_num = request.GET.get('logistic_num', '').strip()
    download_type = request.GET.get('type', 'mgmt').strip()  # mgmt 或 warehouse
    
    if not logistic_num:
        return JsonResponse({'success': False, 'message': _('缺少物流单号')}, status=400)
    
    try:
        # ========== 1. 获取物流单基础信息 ==========
        base_df = DBClient.read_df("""
            SELECT 
                logistic_num,
                date_sent,
                date_eta,
                pallets,
                total_weight,
                price_kg,
                total_price,
                usd_rmb,
                mode,
                seq,
                `by`,
                date_record,
                note,
                CAST(SUBSTRING(seq, 2) AS UNSIGNED) as seq_num
            FROM in_send 
            WHERE logistic_num = :logistic_num
            ORDER BY seq_num DESC
            LIMIT 1
        """, {'logistic_num': logistic_num})
        
        if base_df.empty:
            return JsonResponse({'success': False, 'message': _('发货单不存在')}, status=404)
        
        base_row = base_df.iloc[0]
        date_sent = str(base_row['date_sent']) if base_row['date_sent'] else ''
        
        # ========== 2. 获取货物明细信息 ==========
        detail_df = DBClient.read_df("""
            SELECT 
                seq,
                `by`,
                date,
                note,
                CAST(SUBSTRING(seq, 2) AS UNSIGNED) as seq_num
            FROM in_send_list 
            WHERE logistic_num = :logistic_num
            ORDER BY seq_num DESC
            LIMIT 1
        """, {'logistic_num': logistic_num})
        
        detail_row = detail_df.iloc[0] if not detail_df.empty else None
        
        # ========== 3. 获取货物列表 ==========
        items_df = DBClient.read_df("""
            SELECT 
                po_num,
                po_sku,
                sent_quantity,
                po_price
            FROM in_send_final
            WHERE sent_logistic_num = :logistic_num
            ORDER BY po_num, po_sku
        """, {'logistic_num': logistic_num})
        
        # ========== 4. 计算已全部发货的订单（需要忽略） ==========
        # 获取当前发货单涉及的所有订单号
        po_nums_in_send = items_df['po_num'].unique().tolist() if not items_df.empty else []
        
        fully_shipped_po_nums = set()
        if po_nums_in_send:
            # 获取这些订单在in_po_final中的数据
            po_final_df = DBClient.read_df("""
                SELECT po_num, po_sku, po_quantity, po_price
                FROM in_po_final
                WHERE po_num IN :po_nums
            """, {'po_nums': tuple(po_nums_in_send)})
            
            # 获取这些订单在in_send_final中的发货数据（所有发货单）
            send_final_df = DBClient.read_df("""
                SELECT po_num, po_sku, po_price, SUM(sent_quantity) as total_sent
                FROM in_send_final
                WHERE po_num IN :po_nums
                GROUP BY po_num, po_sku, po_price
            """, {'po_nums': tuple(po_nums_in_send)})
            
            # 构建发货映射
            send_map = {}
            if not send_final_df.empty:
                for _idx, row in send_final_df.iterrows():
                    key = (row['po_num'], row['po_sku'], float(row['po_price']) if row['po_price'] else 0)
                    send_map[key] = int(row['total_sent']) if row['total_sent'] else 0
            
            # 判断每个订单是否全部发货
            for po_num in po_nums_in_send:
                po_items = po_final_df[po_final_df['po_num'] == po_num]
                if po_items.empty:
                    continue
                
                total_ordered = 0
                total_diff = 0
                
                for _idx, item in po_items.iterrows():
                    sku = item['po_sku']
                    qty = int(item['po_quantity']) if item['po_quantity'] else 0
                    price = float(item['po_price']) if item['po_price'] else 0
                    
                    total_ordered += qty
                    key = (po_num, sku, price)
                    sent_qty = send_map.get(key, 0)
                    total_diff += (qty - sent_qty)
                
                # 差值为0表示全部已发货
                if total_ordered > 0 and total_diff == 0:
                    fully_shipped_po_nums.add(po_num)
        
        # ========== 5. 准备计算订货量和已发量的辅助数据 ==========
        # 对于每个(po_num, po_sku, po_price)组合，计算订货量和已发量
        po_nums_list = items_df['po_num'].unique().tolist() if not items_df.empty else []
        
        # 订货量：{(po_num, po_sku, po_price): ordered_qty}
        item_ordered_map = {}
        # 已发量：{(po_num, po_sku, po_price): total_sent}
        item_sent_map = {}
        
        if po_nums_list:
            # 订货量：in_po_final中按(po_num, po_sku, po_price)查询
            ordered_df = DBClient.read_df("""
                SELECT po_num, po_sku, po_price, po_quantity
                FROM in_po_final
                WHERE po_num IN :po_nums
            """, {'po_nums': tuple(po_nums_list)})
            
            if not ordered_df.empty:
                for _idx, row in ordered_df.iterrows():
                    key = (row['po_num'], row['po_sku'], float(row['po_price']) if row['po_price'] else 0)
                    item_ordered_map[key] = int(row['po_quantity']) if row['po_quantity'] else 0
            
            # 已发量：in_send_final中按(po_num, po_sku, po_price)汇总
            sent_df = DBClient.read_df("""
                SELECT po_num, po_sku, po_price, SUM(sent_quantity) as total_sent
                FROM in_send_final
                WHERE po_num IN :po_nums
                GROUP BY po_num, po_sku, po_price
            """, {'po_nums': tuple(po_nums_list)})
            
            if not sent_df.empty:
                for _idx, row in sent_df.iterrows():
                    key = (row['po_num'], row['po_sku'], float(row['po_price']) if row['po_price'] else 0)
                    item_sent_map[key] = int(row['total_sent']) if row['total_sent'] else 0
        
        # ========== 6. 检查是否规整订货量 ==========
        # 获取in_po_final中note包含"物流单据规整操作_"的记录
        def check_is_adjusted(po_num, po_sku, po_price):
            """检查是否存在规整操作记录"""
            note_pattern = f'物流单据规整操作_{logistic_num}'
            adjust_df = DBClient.read_df("""
                SELECT COUNT(*) as cnt
                FROM in_po_final
                WHERE po_num = :po_num 
                  AND po_sku = :po_sku 
                  AND ABS(po_price - :po_price) < 0.001
                  AND po_note LIKE :note_pattern
            """, {
                'po_num': po_num,
                'po_sku': po_sku,
                'po_price': po_price,
                'note_pattern': f'%{note_pattern}%'
            })
            
            if not adjust_df.empty and adjust_df.iloc[0]['cnt'] > 0:
                return _('是')
            return _('否')
        
        # ========== 7. 获取订单日期和货币映射 ==========
        po_date_map = {}
        po_currency_map = {}  # 新增：订单货币映射
        if po_nums_in_send:
            po_date_df = DBClient.read_df("""
                SELECT po_num, po_date
                FROM in_po_final
                WHERE po_num IN :po_nums
                GROUP BY po_num, po_date
            """, {'po_nums': tuple(po_nums_in_send)})
            
            if not po_date_df.empty:
                for _idx, row in po_date_df.iterrows():
                    po_date_map[row['po_num']] = str(row['po_date']) if row['po_date'] else ''
            
            # 获取订单货币信息
            currency_df = DBClient.read_df("""
                SELECT s.po_num, s.cur_currency
                FROM in_po_strategy s
                INNER JOIN (
                    SELECT po_num, MAX(CONCAT(date, seq)) as max_key
                    FROM in_po_strategy
                    WHERE po_num IN :po_nums
                    GROUP BY po_num
                ) latest ON s.po_num = latest.po_num 
                    AND CONCAT(s.date, s.seq) = latest.max_key
            """, {'po_nums': tuple(po_nums_in_send)})
            
            if not currency_df.empty:
                for _idx, row in currency_df.iterrows():
                    po_currency_map[row['po_num']] = row['cur_currency'] or 'RMB'
        
        # ========== 8. 加载模板并填充数据 ==========
        if download_type == 'warehouse':
            template_path = os.path.join(django_settings.BASE_DIR, '..', 'data', 'templates_csv', 'in_send_output_warehouse.xlsx')
        else:
            template_path = os.path.join(django_settings.BASE_DIR, '..', 'data', 'templates_csv', 'in_send_output_mgmt.xlsx')
        
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active
        
        # 辅助函数：安全写入单元格
        def safe_write(cell_ref, value):
            cell = sheet[cell_ref]
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left.value = value
                        return
            cell.value = value
        
        if download_type == 'warehouse':
            # ========== 仓储部门模板填充 ==========
            safe_write('C4', logistic_num)                                    # 物流单号
            safe_write('F4', date_sent)                                       # 发货日期
            safe_write('F6', base_row['seq'] or '-')                         # 最新物流单明细版本
            safe_write('C8', str(base_row['date_eta']) if base_row['date_eta'] else '-')  # 预计到货日期
            safe_write('F8', int(base_row['pallets']) if base_row['pallets'] else 0)      # 托盘数
            
            # 物流单货物信息修订版本
            if detail_row is not None:
                safe_write('F10', detail_row['seq'] or '-')                  # 修订版本
            
            # ========== 填充货物列表（从第13行开始） ==========
            # 先按SKU合并同类项，同时累计订货量和已发量
            sku_merged_data = {}  # {po_sku: {'sent_qty': 0, 'ordered_qty': 0, 'total_sent': 0}}
            
            if not items_df.empty:
                for _idx, item in items_df.iterrows():
                    po_num = item['po_num']
                    po_sku = item['po_sku']
                    po_price = float(item['po_price']) if item['po_price'] else 0
                    sent_qty = int(item['sent_quantity']) if item['sent_quantity'] else 0
                    
                    # 获取该组合的订货量和已发量
                    key = (po_num, po_sku, po_price)
                    ordered_qty = item_ordered_map.get(key, 0)
                    total_sent = item_sent_map.get(key, 0)
                    
                    # 按SKU合并
                    if po_sku not in sku_merged_data:
                        sku_merged_data[po_sku] = {'sent_qty': 0, 'ordered_qty': 0, 'total_sent': 0}
                    sku_merged_data[po_sku]['sent_qty'] += sent_qty
                    sku_merged_data[po_sku]['ordered_qty'] += ordered_qty
                    sku_merged_data[po_sku]['total_sent'] += total_sent
            
            start_row = 13
            current_row = start_row
            
            # 写入合并后的数据
            for po_sku, data in sku_merged_data.items():
                sent_qty = data['sent_qty']
                ordered_qty = data['ordered_qty']
                total_sent = data['total_sent']
                
                # 已发量（不含当前发货）
                already_sent = total_sent - sent_qty
                # 未发量 = 订货量 - 已发量(不含当前) - 发货量(当前)
                unshipped_qty = ordered_qty - already_sent - sent_qty
                
                # 写入Excel (仓储部门模板列)
                sheet.cell(row=current_row, column=2).value = po_sku       # B列: SKU
                sheet.cell(row=current_row, column=3).value = ordered_qty  # C列: 订货量
                sheet.cell(row=current_row, column=4).value = already_sent # D列: 已发量（不含当前发货）
                sheet.cell(row=current_row, column=5).value = sent_qty     # E列: 发货量
                sheet.cell(row=current_row, column=6).value = unshipped_qty  # F列: 未发量
                
                current_row += 1
            
            # 文件名
            filename = f"{logistic_num}_{date_sent}_warehouse_current.xlsx"
        
        else:
            # ========== 管理部门模板填充 ==========
            safe_write('C4', logistic_num)                                    # 物流单号
            safe_write('J4', date_sent)                                       # 发货日期
            safe_write('J6', base_row['seq'] or '-')                         # 最新物流单明细版本
            safe_write('C8', base_row['by'] or '-')                          # 操作人
            safe_write('J8', str(base_row['date_record']) if base_row['date_record'] else '-')  # 修订日期
            safe_write('C10', base_row['note'] or '-')                       # 备注
            safe_write('C12', str(base_row['date_eta']) if base_row['date_eta'] else '-')       # 预计到货日期
            safe_write('F12', int(base_row['pallets']) if base_row['pallets'] else 0)           # 托盘数
            safe_write('J12', float(base_row['total_weight']) if base_row['total_weight'] else 0)  # 发货总重量(KG)
            safe_write('C14', float(base_row['price_kg']) if base_row['price_kg'] else 0)       # 物流单价(RMB/KG)
            safe_write('F14', float(base_row['total_price']) if base_row['total_price'] else 0) # 物流总价(RMB)
            safe_write('J14', float(base_row['usd_rmb']) if base_row['usd_rmb'] else 0)         # 结算汇率
            
            # 物流单货物信息
            if detail_row is not None:
                safe_write('J16', detail_row['seq'] or '-')                  # 修订版本
                safe_write('C18', detail_row['by'] or '-')                   # 操作人
                safe_write('J18', str(detail_row['date']) if detail_row['date'] else '-')  # 修订日期
                safe_write('C20', detail_row['note'] or '-')                 # 备注
            
            # ========== 填充货物列表（从第23行开始） ==========
            start_row = 23
            current_row = start_row
            
            if not items_df.empty:
                for _idx, item in items_df.iterrows():
                    po_num = item['po_num']
                    po_sku = item['po_sku']
                    po_price = float(item['po_price']) if item['po_price'] else 0
                    sent_qty = int(item['sent_quantity']) if item['sent_quantity'] else 0

                    
                    # 使用(po_num, po_sku, po_price)组合获取数据
                    key = (po_num, po_sku, po_price)
                    
                    # 订单日期
                    po_date = po_date_map.get(po_num, '')
                    # 订货量
                    ordered_qty = item_ordered_map.get(key, 0)
                    # 已发量（包含当前发货）
                    total_sent = item_sent_map.get(key, 0)
                    # 已发量（不含当前发货）
                    already_sent = total_sent - sent_qty
                    # 未发量 = 订货量 - 已发量(不含当前) - 发货量(当前)
                    unshipped_qty = ordered_qty - already_sent - sent_qty
                    # 是否规整订货量
                    is_adjusted = check_is_adjusted(po_num, po_sku, po_price)
                    
                    # 写入Excel (管理部门模板列)
                    sheet.cell(row=current_row, column=2).value = po_date      # B列: 订单日期
                    sheet.cell(row=current_row, column=3).value = po_num       # C列: 订单号
                    sheet.cell(row=current_row, column=4).value = po_sku       # D列: SKU
                    # E列: 单价（含货币）
                    currency = po_currency_map.get(po_num, 'RMB')
                    price_display = f"{currency} {po_price}"
                    sheet.cell(row=current_row, column=5).value = price_display
                    sheet.cell(row=current_row, column=6).value = ordered_qty  # F列: 订货量
                    sheet.cell(row=current_row, column=7).value = already_sent # G列: 已发量（不含当前发货）
                    sheet.cell(row=current_row, column=8).value = sent_qty     # H列: 发货量
                    sheet.cell(row=current_row, column=9).value = unshipped_qty  # I列: 未发量
                    sheet.cell(row=current_row, column=10).value = is_adjusted # J列: 是否规整订货量
                    
                    # 条件样式：未发量≠0 且 是否规整订货量为"是" 时，设置橘黄色背景
                    if unshipped_qty != 0 and is_adjusted == _('是'):
                        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                        sheet.cell(row=current_row, column=9).fill = orange_fill
                    
                    current_row += 1
            
            # 文件名
            filename = f"{logistic_num}_{date_sent}_mgmt_current.xlsx"
        
        # ========== 10. 保存并返回文件 ==========
        import tempfile
        
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        try:
            os.close(temp_fd)
            wb.save(temp_path)
            
            with open(temp_path, 'rb') as f:
                file_content = f.read()
        finally:
            try:
                os.unlink(temp_path)
            except:
                pass
        
        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.exception("操作失败")
        return JsonResponse({
            'success': False,
            'message': _('生成发货单文件失败: {error}').format(error=str(e))
        }, status=500)


