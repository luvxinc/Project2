"""
新建发货单 - 模板相关API
包含：生成模板数据、下载模板、验证Excel

[P0-2 Fix] 统一使用 hub.check_perm
[P1-2 Fix] 使用 logging 替代 traceback.print_exc
"""
import os
import json
import logging
from datetime import datetime
from openpyxl import load_workbook

from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET, require_POST
from django.conf import settings as django_settings

from ..hub import check_perm
from core.components.db.client import DBClient
from django.utils.translation import gettext as _

logger = logging.getLogger(__name__)


@login_required(login_url='web_ui:login')
@require_GET
def generate_template_data_api(request):
    """
    生成模板数据（用于流程2验证比对）
    Permission: module.purchase.send.add
    
    GET 参数:
        date_sent: 发货日期，用于筛选 po_date <= date_sent 的订单
    
    返回与下载模板相同的数据结构，但不生成Excel文件
    前端将此数据保存在内存中，用于后续验证比对
    """
    if not check_perm(request.user, 'module.purchase.send.add'):
        return JsonResponse({'success': False, 'message': _('无权限')}, status=403)
    
    # 获取发货日期参数
    date_sent = request.GET.get('date_sent', '').strip()
    if not date_sent:
        return JsonResponse({
            'success': False, 
            'message': _('请先选择发货日期'),
            'error_type': 'no_date'
        })
    
    try:
        # 读取 in_po_final 表数据，筛选 po_date <= date_sent
        po_final_df = DBClient.read_df("""
            SELECT po_date, po_num, po_sku, po_quantity, po_price
            FROM in_po_final
            WHERE po_date <= :date_sent
            ORDER BY po_num, po_sku, po_price DESC
        """, {'date_sent': date_sent})
        
        if po_final_df.empty:
            return JsonResponse({
                'success': False, 
                'message': _('该发货日期({date_sent})前没有订单，请先去建立订单').format(date_sent=date_sent),
                'error_type': 'no_orders'
            })
        
        # 读取 in_send_final 表数据（按 po_num, po_sku, po_price 分组）
        send_final_df = DBClient.read_df("""
            SELECT po_num, po_sku, po_price, SUM(sent_quantity) as send_quantity
            FROM in_send_final
            GROUP BY po_num, po_sku, po_price
        """)
        
        # 创建 sent 数据的字典 {(po_num, po_sku, po_price): send_quantity}
        sent_dict = {}
        if not send_final_df.empty:
            for _idx, row in send_final_df.iterrows():
                key = (row['po_num'], row['po_sku'], float(row['po_price']) if row['po_price'] else 0)
                sent_dict[key] = int(row['send_quantity']) if row['send_quantity'] else 0
        
        # ========== 计算每个订单是否全部已发货 ==========
        po_nums_in_data = po_final_df['po_num'].unique().tolist()
        fully_shipped_po_nums = set()
        
        for po_num in po_nums_in_data:
            po_items = po_final_df[po_final_df['po_num'] == po_num]
            total_diff = 0
            total_ordered = 0
            
            for _idx, item in po_items.iterrows():
                sku = item['po_sku']
                qty = int(item['po_quantity']) if item['po_quantity'] else 0
                price = float(item['po_price']) if item['po_price'] else 0
                
                total_ordered += qty
                
                key = (po_num, sku, price)
                sent_qty = sent_dict.get(key, 0)
                total_diff += (qty - sent_qty)
            
            if total_ordered > 0 and total_diff == 0:
                fully_shipped_po_nums.add(po_num)
        
        # ========== 创建按(po_num, po_sku)汇总的发货量字典 ==========
        sent_by_sku_dict = {}
        if not send_final_df.empty:
            for _idx, row in send_final_df.iterrows():
                key = (row['po_num'], row['po_sku'])
                sent_by_sku_dict[key] = sent_by_sku_dict.get(key, 0) + (int(row['send_quantity']) if row['send_quantity'] else 0)
        
        # 统计每个(po_num, po_sku)下有多少条记录（用于价格位列）
        sku_count = {}
        if not po_final_df.empty:
            for _idx, po_row in po_final_df.iterrows():
                key = (po_row['po_num'], po_row['po_sku'])
                sku_count[key] = sku_count.get(key, 0) + 1
        
        sku_current_rank = {}
        
        # 生成模板数据
        items = []
        for _idx, po_row in po_final_df.iterrows():
            po_num = po_row['po_num']
            po_sku = po_row['po_sku']
            po_quantity = int(po_row['po_quantity']) if po_row['po_quantity'] else 0
            po_date = str(po_row['po_date']) if po_row['po_date'] else ''
            
            if po_num in fully_shipped_po_nums:
                continue
            
            key = (po_num, po_sku)
            send_quantity = sent_by_sku_dict.get(key, 0)
            remaining = po_quantity - send_quantity
            
            if remaining <= 0:
                continue
            
            price_rank_str = ''
            if sku_count.get(key, 1) > 1:
                sku_current_rank[key] = sku_current_rank.get(key, 0) + 1
                price_rank_str = _('价格位列{rank}').format(rank=sku_current_rank[key])
            
            items.append({
                'row': len(items) + 9,
                'B': po_date,
                'C': po_num,
                'D': po_sku,
                'E': po_quantity,
                'F': send_quantity,
                'G': remaining,
                'J': price_rank_str
            })
        
        if not items:
            return JsonResponse({
                'success': False, 
                'message': _('该发货日期({date})前的订单已全部发货完成').format(date=date_sent),
                'error_type': 'all_shipped'
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'items': items,
                'date_sent': date_sent,
                'generated_at': datetime.now().isoformat()
            }
        })
        
    except Exception as e:
        logger.exception("生成模板数据失败")
        return JsonResponse({'success': False, 'message': _('生成模板数据失败: {error}').format(error=str(e))}, status=500)


@login_required(login_url='web_ui:login')
@require_POST
def download_send_template_api(request):
    """
    下载发货模板Excel文件
    Permission: module.purchase.send.add
    """
    if not check_perm(request.user, 'module.purchase.send.add'):
        return JsonResponse({'success': False, 'message': _('无权限')}, status=403)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': _('数据格式错误')}, status=400)
    
    date_sent = data.get('date_sent', '').strip()
    items = data.get('items', [])
    
    if not date_sent:
        return JsonResponse({'success': False, 'message': _('请先选择发货日期')}, status=400)
    
    if not items:
        return JsonResponse({'success': False, 'message': _('没有可发货的订单数据')}, status=400)
    
    try:
        template_path = os.path.join(django_settings.BASE_DIR, '..', 'data', 'templates_csv', 'in_send_upload.xlsx')
        
        if not os.path.exists(template_path):
            return JsonResponse({'success': False, 'message': _('模板文件不存在')}, status=404)
        
        wb = load_workbook(template_path)
        ws = wb.active
        
        ws['C2'] = date_sent
        
        row_num = 9
        for item in items:
            ws[f'B{row_num}'] = item.get('B', '')
            ws[f'C{row_num}'] = item.get('C', '')
            ws[f'D{row_num}'] = item.get('D', '')
            ws[f'E{row_num}'] = item.get('E', 0)
            ws[f'F{row_num}'] = item.get('F', 0)
            ws[f'G{row_num}'] = item.get('G', 0)
            ws[f'J{row_num}'] = item.get('J', '')
            row_num += 1
        
        from openpyxl.styles import Protection
        from openpyxl.worksheet.datavalidation import DataValidation
        
        ws['I4'] = date_sent
        
        unlocked_cells = ['C4', 'F4', 'C6', 'F6', 'I6']
        for cell_ref in unlocked_cells:
            ws[cell_ref].protection = Protection(locked=False)
        
        dv = DataValidation(
            type="list",
            formula1='"是,否"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="请从下拉列表中选择：是 或 否"
        )
        dv.prompt = "请选择是否规整订货量"
        dv.promptTitle = "规整选项"
        ws.add_data_validation(dv)
        
        data_start_row = 9
        data_end_row = row_num - 1
        
        for r in range(data_start_row, data_end_row + 1):
            sku_val = ws[f'D{r}'].value
            if sku_val and str(sku_val).strip():
                ws[f'H{r}'].protection = Protection(locked=False)
                ws[f'I{r}'].protection = Protection(locked=False)
                dv.add(ws[f'I{r}'])
        
        ws.protection.sheet = True
        ws.protection.password = '1522'
        ws.protection.enable()
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="shipping_form_{date_sent}.xlsx"'
        
        return response
        
    except Exception as e:
        logger.exception("生成模板失败")
        return JsonResponse({'success': False, 'message': _('生成模板失败: {error}').format(error=str(e))}, status=500)


@login_required(login_url='web_ui:login')
@require_POST
def validate_send_excel_api(request):
    """
    验证上传的发货模板Excel文件
    Permission: module.purchase.send.add
    """
    if not check_perm(request.user, 'module.purchase.send.add'):
        return JsonResponse({'success': False, 'message': _('无权限')}, status=403)
    
    try:
        from io import BytesIO
        import openpyxl
        
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({'success': False, 'message': _('请上传文件')})
        
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'message': _('仅支持 .xlsx 或 .xls 格式的Excel文件')})
        
        try:
            uploaded_wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()), data_only=True)
            uploaded_ws = uploaded_wb.active
        except Exception as e:
            return JsonResponse({'success': False, 'message': _('无法解析Excel文件: {error}').format(error=str(e))})
        
        # 规则1: 数据一致性比对
        template_data_str = request.POST.get('template_data', '')
        
        if not template_data_str:
            return JsonResponse({
                'success': False,
                'error_type': 'mismatch',
                'title': '文件上传错误',
                'message': _('无法验证上传文件，请返回第一步重新下载模板'),
                'details': ['请点击上方步骤条返回"发货说明"步骤', '下载最新发货模板', '填写后重新上传']
            })
        
        differences = []
        
        try:
            template_data = json.loads(template_data_str)
            template_items = template_data.get('items', [])
            
            if not template_items:
                return JsonResponse({
                    'success': False,
                    'error_type': 'mismatch',
                    'title': _('文件上传错误'),
                    'message': _('当前无可发货的订单数据，请确认后重试：'),
                    'details': [_('请确认系统中存在待发货订单'), _('返回第一步重新下载模板')]
                })
            
            for item in template_items:
                row = item.get('row', 0)
                if row < 9:
                    continue
                
                col_mapping = {
                    'B': item.get('B', ''),
                    'C': item.get('C', ''),
                    'D': item.get('D', ''),
                    'E': item.get('E', ''),
                    'G': item.get('G', '')
                }
                
                for col, expected_val in col_mapping.items():
                    if expected_val is None or str(expected_val).strip() == '':
                        continue
                    
                    uploaded_val = uploaded_ws[f'{col}{row}'].value
                    expected_str = str(expected_val).strip()
                    uploaded_str = str(uploaded_val).strip() if uploaded_val is not None else ''
                    
                    if expected_str != uploaded_str:
                        differences.append(f'{col}{row}: 期望"{expected_str}", 实际"{uploaded_str}"')
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error_type': 'mismatch',
                'title': _('文件上传错误'),
                'message': _('验证数据异常，请返回第一步重新开始：'),
                'details': [_('请点击上方步骤条返回"发货说明"步骤'), _('下载最新发货模板'), _('填写后重新上传')]
            })
        
        if differences:
            return JsonResponse({
                'success': False,
                'error_type': 'mismatch',
                'title': _('文件上传错误'),
                'message': _('上传的文件与当前发货模板数据不一致，请确保上传的是最新下载的模板：'),
                'details': differences[:10] + ([_('...还有更多不一致项')] if len(differences) > 10 else [])
            })
        
        errors = []
        
        # 规则2: 检查物流参数单元格
        cell_labels = {
            'C4': '物流单号',
            'F4': '托盘数',
            'I4': '发货日期',
            'C6': '物流单价',
            'F6': '发货总重量',
            'I6': '预计到货日期'
        }
        
        required_cells = ['C4', 'F4', 'I4', 'C6', 'F6', 'I6']
        for cell_ref in required_cells:
            cell_value = uploaded_ws[cell_ref].value
            if cell_value is None or str(cell_value).strip() == '':
                label = cell_labels.get(cell_ref, cell_ref)
                errors.append(f'{label} ({cell_ref})')
        
        if errors:
            return JsonResponse({
                'success': False, 
                'error_type': 'incomplete',
                'title': '上传文件不完整',
                'message': _('以下物流参数未填写'),
                'details': errors
            })
        
        # 规则3: 检查发货明细
        row = 9
        item_errors = []
        max_check_rows = 500
        while row < 9 + max_check_rows:
            sku = uploaded_ws[f'D{row}'].value
            if sku is None or str(sku).strip() == '':
                row += 1
                continue
            
            sku_str = str(sku).strip()
            h_value = uploaded_ws[f'H{row}'].value
            i_value = uploaded_ws[f'I{row}'].value
            
            if h_value is not None and str(h_value).strip() != '':
                try:
                    h_int = int(float(str(h_value)))
                    if h_int <= 0:
                        item_errors.append(f'第{row}行 {sku_str} - 发货量必须大于0 (H{row})')
                except (ValueError, TypeError):
                    item_errors.append(f'第{row}行 {sku_str} - 发货量必须是整数 (H{row})')
            
            if i_value is not None and str(i_value).strip() != '':
                if str(i_value).strip() not in ['是', '否']:
                    item_errors.append(f'第{row}行 {sku_str} - 是否规整订货量必须为"是"或"否" (I{row})')
            
            row += 1
            
            if row > 9:
                empty_count = 0
                for check_row in range(row, min(row + 10, 9 + max_check_rows)):
                    check_val = uploaded_ws[f'D{check_row}'].value
                    if check_val is None or str(check_val).strip() == '':
                        empty_count += 1
                    else:
                        break
                if empty_count >= 10:
                    break
        
        if item_errors:
            return JsonResponse({
                'success': False,
                'error_type': 'incomplete',
                'title': _('上传文件不完整'),
                'message': _('以下发货明细未填写：'),
                'details': item_errors[:10] + ([_('...还有更多未填写项')] if len(item_errors) > 10 else [])
            })
        
        # 日期格式智能转换
        def parse_date_to_ymd(value):
            if value is None:
                return ''
            
            from datetime import datetime as dt, date
            if isinstance(value, (dt, date)):
                return value.strftime('%Y-%m-%d')
            
            date_str = str(value).strip()
            if not date_str:
                return ''
            
            formats_to_try = [
                '%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y',
                '%Y.%m.%d', '%d-%m-%Y', '%Y年%m月%d日',
                '%d-%b-%Y', '%d-%b-%y', '%b %d, %Y', '%b %d %Y',
                '%d %b %Y', '%d %b, %Y', '%B %d, %Y', '%B %d %Y',
                '%d %B %Y', '%Y-%b-%d', '%Y %b %d',
            ]
            
            for fmt in formats_to_try:
                try:
                    parsed = dt.strptime(date_str, fmt)
                    return parsed.strftime('%Y-%m-%d')
                except ValueError:
                    continue
            
            try:
                excel_serial = float(date_str)
                if 1 < excel_serial < 100000:
                    from datetime import timedelta
                    base_date = dt(1899, 12, 30)
                    result_date = base_date + timedelta(days=excel_serial)
                    return result_date.strftime('%Y-%m-%d')
            except (ValueError, TypeError):
                pass
            
            return date_str
        
        # 提取物流参数
        logistics_data = {
            'date_sent': parse_date_to_ymd(uploaded_ws['I4'].value),
            'date_eta': parse_date_to_ymd(uploaded_ws['I6'].value),
            'logistic_num': str(uploaded_ws['C4'].value) if uploaded_ws['C4'].value else '',
            'pallet_count': int(uploaded_ws['F4'].value) if uploaded_ws['F4'].value else 0,
            'total_weight': float(uploaded_ws['F6'].value) if uploaded_ws['F6'].value else 0,
            'price_kg': float(uploaded_ws['C6'].value) if uploaded_ws['C6'].value else 0
        }
        
        items_data = []
        row = 9
        while True:
            b_value = uploaded_ws[f'B{row}'].value
            if b_value is None or str(b_value).strip() == '':
                break
            
            items_data.append({
                'row': row,
                'po_date': str(b_value),
                'po_num': str(uploaded_ws[f'C{row}'].value) if uploaded_ws[f'C{row}'].value else '',
                'po_sku': str(uploaded_ws[f'D{row}'].value) if uploaded_ws[f'D{row}'].value else '',
                'po_quantity': int(uploaded_ws[f'E{row}'].value) if uploaded_ws[f'E{row}'].value else 0,
                'send_quantity': int(uploaded_ws[f'H{row}'].value) if uploaded_ws[f'H{row}'].value else 0,
                'is_rounded': uploaded_ws[f'I{row}'].value is not None and str(uploaded_ws[f'I{row}'].value).strip() == '是'
            })
            row += 1
        
        return JsonResponse({
            'success': True,
            'message': f'文件验证成功！共{len(items_data)}条发货记录',
            'data': {
                'logistics': logistics_data,
                'items': items_data
            }
        })
        
    except Exception as e:
        logger.exception("验证失败")
        return JsonResponse({'success': False, 'message': _('验证失败: {error}').format(error=str(e))}, status=500)

